#!/usr/bin/env python3
"""Index R replica plotting scripts into reusable code-pattern metadata.

The index is intentionally structural rather than a code copier: it records
what the scripts read, which plotting layers they use, what data roles appear
in the inputs, and whether the case should become a production recipe,
template candidate, specialized reference, or caution example.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable
from xml.etree import ElementTree as ET


SCRIPT_EXTS = {".r", ".rmd"}
DATA_EXTS = {".csv", ".tsv", ".txt", ".xls", ".xlsx", ".nwk", ".geojson", ".json"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".pdf", ".svg", ".tif", ".tiff", ".ai"}

PACKAGE_RE = re.compile(r"\b(?:library|require)\s*\(\s*['\"]?([A-Za-z0-9_.]+)['\"]?", re.I)
NAMESPACE_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9_.]+)::")
CALL_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9_.]+)\s*\(")
AES_RE = re.compile(r"\baes(?:_string)?\s*\((.*?)\)", re.S)
READ_RE = re.compile(
    r"\b(read\.csv|read\.table|read\.delim|readr::read_csv|readr::read_tsv|"
    r"readxl::read_excel|read_excel|fread|data\.table::fread|readRDS)\s*\((.*?)\)",
    re.S,
)
OUTPUT_RE = re.compile(r"\b(ggsave|pdf|png|svg|tiff|jpeg|cairo_pdf)\s*\(", re.I)
HARDCODED_PATH_RE = re.compile(r"(/Users/|~/|[A-Za-z]:\\\\)")

ROLE_PATTERNS = {
    "sample": ("sample", "sampleid", "accession", "strain", "line", "id", "taxa"),
    "group": ("group", "condition", "treatment", "type", "class", "cluster", "population"),
    "metric": ("metric", "trait", "variable", "index", "term", "pathway"),
    "value": ("value", "abundance", "expression", "score", "mean", "rate", "ratio", "fraction"),
    "error": ("se", "sd", "stderr", "ci", "lower", "upper", "error", "sem"),
    "feature": ("gene", "feature", "otu", "asv", "protein", "transcript"),
    "pvalue": ("pvalue", "p.value", "p_val", "padj", "qvalue", "fdr"),
    "effect": ("logfc", "log2fc", "foldchange", "estimate", "effect"),
    "coordinate": ("chr", "chrom", "chromosome", "pos", "position", "start", "end"),
    "ordination": ("pc1", "pc2", "pcoa", "nmds", "umap", "tsne"),
    "network": ("source", "target", "from", "to", "node", "edge"),
    "spatial": ("lon", "long", "longitude", "lat", "latitude", "geometry"),
    "tree": ("newick", "nwk", "tree", "tip", "clade"),
}

SPECIALIZED_PACKAGES = {
    "circlize",
    "ComplexHeatmap",
    "ggtree",
    "ape",
    "sf",
    "maps",
    "mapdata",
    "networkD3",
    "ggalluvial",
    "ComplexUpset",
    "UpSetR",
    "scatterpie",
    "ggtern",
    "ggraph",
    "igraph",
    "ggforce",
}

COMMON_PACKAGES = {
    "ggplot2",
    "dplyr",
    "tidyr",
    "readr",
    "readxl",
    "data.table",
    "tibble",
    "forcats",
    "stringr",
    "RColorBrewer",
    "scales",
    "cowplot",
    "patchwork",
    "gridExtra",
    "ggrepel",
}


def read_text(path: Path) -> str:
    for encoding in ("utf-8", "utf-8-sig", "gb18030", "latin1"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        except OSError:
            return ""
    return path.read_text(errors="ignore")


def strip_comments(text: str) -> str:
    out = []
    for line in text.splitlines():
        if "#" in line:
            line = line.split("#", 1)[0]
        out.append(line)
    return "\n".join(out)


def unique_sorted(values: Iterable[str]) -> list[str]:
    return sorted({str(v) for v in values if str(v)})


def split_top_level_args(arg_text: str) -> list[str]:
    args: list[str] = []
    depth = 0
    quote = None
    start = 0
    for i, ch in enumerate(arg_text):
        if quote:
            if ch == quote and (i == 0 or arg_text[i - 1] != "\\"):
                quote = None
            continue
        if ch in {"'", '"'}:
            quote = ch
        elif ch in "([{":
            depth += 1
        elif ch in ")]}":
            depth = max(0, depth - 1)
        elif ch == "," and depth == 0:
            args.append(arg_text[start:i].strip())
            start = i + 1
    tail = arg_text[start:].strip()
    if tail:
        args.append(tail)
    return args


def parse_aes_roles(text: str) -> dict[str, list[str]]:
    roles: dict[str, set[str]] = defaultdict(set)
    for match in AES_RE.finditer(text):
        for arg in split_top_level_args(match.group(1)):
            if "=" in arg:
                lhs, rhs = arg.split("=", 1)
                lhs = lhs.strip()
                rhs = rhs.strip()
            else:
                lhs, rhs = "implicit", arg.strip()
            rhs = re.sub(r"\.data\[\[['\"]([^'\"]+)['\"]\]\]", r"\1", rhs)
            rhs = re.sub(r"`([^`]+)`", r"\1", rhs)
            rhs = re.sub(r"[^A-Za-z0-9_.]+", " ", rhs).strip()
            tokens = [tok for tok in rhs.split() if not tok.replace(".", "", 1).isdigit()]
            for token in tokens[:2]:
                if token and token not in {"TRUE", "FALSE", "NA", "NULL"}:
                    roles[lhs].add(token)
    return {k: unique_sorted(v) for k, v in roles.items()}


def read_csv_header(path: Path) -> list[str]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            dialect = csv.Sniffer().sniff(handle.read(4096))
            handle.seek(0)
            reader = csv.reader(handle, dialect)
            return next(reader, [])
    except Exception:
        try:
            with path.open("r", encoding="gb18030", newline="") as handle:
                reader = csv.reader(handle)
                return next(reader, [])
        except Exception:
            return []


def read_xlsx_header(path: Path) -> list[str]:
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(max_row=1, values_only=True), ())
        return [str(x) for x in first_row if x is not None]
    except Exception:
        pass
    try:
        with zipfile.ZipFile(path) as zf:
            shared = []
            if "xl/sharedStrings.xml" in zf.namelist():
                root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                for si in root.findall("a:si", ns):
                    shared.append("".join(t.text or "" for t in si.findall(".//a:t", ns)))
            sheet_name = next(x for x in zf.namelist() if x.startswith("xl/worksheets/sheet"))
            root = ET.fromstring(zf.read(sheet_name))
            ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            row = root.find(".//a:sheetData/a:row", ns)
            if row is None:
                return []
            values = []
            for cell in row.findall("a:c", ns):
                value = cell.find("a:v", ns)
                if value is None:
                    continue
                text = value.text or ""
                if cell.attrib.get("t") == "s" and text.isdigit() and int(text) < len(shared):
                    text = shared[int(text)]
                values.append(text)
            return values
    except Exception:
        return []


def headers_for_file(path: Path) -> list[str]:
    ext = path.suffix.lower()
    if ext in {".csv", ".tsv", ".txt"}:
        return read_csv_header(path)
    if ext == ".xlsx":
        return read_xlsx_header(path)
    return []


def infer_roles(columns: Iterable[str]) -> dict[str, list[str]]:
    roles: dict[str, list[str]] = {}
    for col in columns:
        key = re.sub(r"[^a-z0-9]+", "", col.lower())
        for role, patterns in ROLE_PATTERNS.items():
            if any(pat.replace(".", "") in key for pat in patterns):
                roles.setdefault(role, []).append(col)
    return {k: unique_sorted(v) for k, v in roles.items()}


def infer_family(case_name: str, text: str, geoms: list[str], packages: list[str], data_exts: list[str]) -> str:
    lower = " ".join([case_name.lower(), text.lower(), " ".join(geoms), " ".join(packages), " ".join(data_exts)])
    if "manhattan" in lower or {"chr", "position"} <= set(lower.split()):
        return "Manhattan / genomewide"
    if "volcano" in lower:
        return "volcano / enrichment"
    if "gsea" in lower or "enrich" in lower:
        return "enrichment / GSEA"
    if "pcoa" in lower or "nmds" in lower or "pca" in lower or "umap" in lower:
        return "PCA / PCoA / ordination"
    if "heatmap" in lower or "geom_tile" in geoms or "complexheatmap" in lower:
        return "heatmap / matrix"
    if "upset" in lower or "venn" in lower:
        return "upset / set plot"
    if "circos" in lower or "circlize" in lower or "chord" in lower:
        return "circos / chord / circular"
    if "sankey" in lower or "network" in lower or "igraph" in lower:
        return "network / sankey"
    if "map" in lower or "sf" in packages or "geojson" in data_exts:
        return "map / spatial"
    if "tree" in lower or "ggtree" in lower or "nwk" in data_exts:
        return "phylogenetic tree"
    if "ridge" in lower or "density" in lower:
        return "ridgeline / density"
    if "raincloud" in lower or "violin" in lower or "geom_violin" in geoms:
        return "raincloud / violin / jitter"
    if "boxplot" in lower or "geom_boxplot" in geoms:
        return "boxplot / jitter"
    if "forest" in lower or "effect" in lower:
        return "forest / effect-size"
    if "lollipop" in lower or ("geom_segment" in geoms and "geom_point" in geoms):
        return "lollipop / dumbbell / dotplot"
    if "bar" in lower or "geom_col" in geoms or "geom_bar" in geoms:
        return "grouped bar / stacked bar / errorbar"
    if "scatter" in lower or "geom_point" in geoms:
        return "scatter / regression"
    return "multi-panel manuscript / other"


def classify_suitability(packages: list[str], family: str, text: str, geoms: list[str]) -> tuple[str, list[str]]:
    risks: list[str] = []
    package_set = set(packages)
    if HARDCODED_PATH_RE.search(text):
        risks.append("hardcoded_local_path")
    if len(package_set - COMMON_PACKAGES) >= 8:
        risks.append("many_dependencies")
    if package_set & SPECIALIZED_PACKAGES:
        risks.append("specialized_package")
    if "annotation_custom" in text or "grid::" in text or "grob" in text.lower():
        risks.append("case_specific_grob_layout")
    if "theme_void" in text.lower() or "polar" in text.lower() or "coord_polar" in text:
        risks.append("decorative_or_polar_layout")
    if family in {
        "circos / chord / circular",
        "network / sankey",
        "map / spatial",
        "phylogenetic tree",
        "upset / set plot",
    }:
        return "specialized_reference", risks
    if "decorative_or_polar_layout" in risks or "case_specific_grob_layout" in risks:
        return "decorative_or_case_specific", risks
    if "many_dependencies" in risks:
        return "specialized_reference", risks
    if family in {
        "grouped bar / stacked bar / errorbar",
        "raincloud / violin / jitter",
        "scatter / regression",
        "heatmap / matrix",
        "PCA / PCoA / ordination",
        "volcano / enrichment",
        "enrichment / GSEA",
        "forest / effect-size",
        "lollipop / dumbbell / dotplot",
        "ridgeline / density",
    }:
        if len(geoms) >= 2:
            return "template_candidate", risks
        return "production_recipe", risks
    return "production_recipe", risks


def script_metadata(script: Path, case_dir: Path, library: Path, archive_root: Path | None) -> dict:
    text_raw = read_text(script)
    text = strip_comments(text_raw)
    packages = unique_sorted(PACKAGE_RE.findall(text) + NAMESPACE_RE.findall(text))
    calls = CALL_RE.findall(text)
    geoms = unique_sorted(x for x in calls if x.startswith("geom_"))
    stats = unique_sorted(x for x in calls if x.startswith("stat_"))
    scales = unique_sorted(x for x in calls if x.startswith("scale_"))
    themes = unique_sorted(x for x in calls if x.startswith("theme"))
    facets = unique_sorted(x for x in calls if x.startswith("facet_"))
    coords = unique_sorted(x for x in calls if x.startswith("coord_"))
    outputs = unique_sorted(OUTPUT_RE.findall(text))
    reads = []
    for reader, args in READ_RE.findall(text):
        reads.append({"reader": reader, "argument_preview": re.sub(r"\s+", " ", args.strip())[:160]})
    data_files = [p for p in case_dir.rglob("*") if p.is_file() and p.suffix.lower() in DATA_EXTS]
    image_files = [p for p in case_dir.rglob("*") if p.is_file() and p.suffix.lower() in IMAGE_EXTS]
    headers = {}
    all_columns: list[str] = []
    for data_file in data_files[:8]:
        header = headers_for_file(data_file)
        if header:
            rel = data_file.relative_to(library).as_posix()
            headers[rel] = header[:80]
            all_columns.extend(header)
    aes_roles = parse_aes_roles(text)
    data_roles = infer_roles(all_columns + [x for values in aes_roles.values() for x in values])
    data_exts = unique_sorted(p.suffix.lower().lstrip(".") for p in data_files)
    output_exts = unique_sorted(p.suffix.lower().lstrip(".") for p in image_files)
    case_name = case_dir.name
    family = infer_family(case_name, text, geoms, packages, data_exts)
    suitability, risks = classify_suitability(packages, family, text, geoms)
    archive_rel = None
    if archive_root is not None:
        try:
            archive_rel = (archive_root / script.relative_to(library)).as_posix()
        except ValueError:
            archive_rel = None
    return {
        "case": case_name,
        "script": script.relative_to(library).as_posix(),
        "archive_script": archive_rel,
        "language": "Rmd" if script.suffix.lower() == ".rmd" else "R",
        "figure_family": family,
        "suitability": suitability,
        "packages": packages,
        "data_readers": reads,
        "data_file_count": len(data_files),
        "data_file_types": data_exts,
        "image_output_count": len(image_files),
        "image_output_types": output_exts,
        "ggplot_layers": {
            "geoms": geoms,
            "stats": stats,
            "scales": scales,
            "themes": themes,
            "facets": facets,
            "coords": coords,
        },
        "aes_roles": aes_roles,
        "input_schema_roles": data_roles,
        "sample_headers": headers,
        "output_calls": outputs,
        "risks": risks,
    }


def write_markdown(records: list[dict], out: Path, library: Path) -> None:
    counts = Counter(r["suitability"] for r in records)
    families = Counter(r["figure_family"] for r in records)
    lines = [
        "# Replica Code Pattern Index",
        "",
        f"Library: `{library}`",
        f"Scripts indexed: {len(records)}",
        "",
        "## Suitability Summary",
        "",
    ]
    for key, value in sorted(counts.items()):
        lines.append(f"- `{key}`: {value}")
    lines += ["", "## Figure Family Summary", ""]
    for key, value in families.most_common():
        lines.append(f"- {key}: {value}")
    lines += [
        "",
        "## Script-Level Index",
        "",
        "| Case | Script | Family | Suitability | Packages | Geoms | Input roles | Risks |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for r in records:
        packages = ", ".join(r["packages"][:8])
        if len(r["packages"]) > 8:
            packages += ", ..."
        geoms = ", ".join(r["ggplot_layers"]["geoms"][:8])
        roles = ", ".join(sorted(r["input_schema_roles"].keys()))
        risks = ", ".join(r["risks"])
        lines.append(
            "| {case} | `{script}` | {family} | `{suitability}` | {packages} | {geoms} | {roles} | {risks} |".format(
                case=r["case"].replace("|", "/"),
                script=r["script"].replace("|", "/"),
                family=r["figure_family"].replace("|", "/"),
                suitability=r["suitability"],
                packages=packages.replace("|", "/"),
                geoms=geoms.replace("|", "/"),
                roles=roles.replace("|", "/"),
                risks=risks.replace("|", "/"),
            )
        )
    lines += [
        "",
        "## How To Use This Index",
        "",
        "- Use `production_recipe` and `template_candidate` cases as source evidence for reusable code recipes.",
        "- Use `specialized_reference` cases to learn data requirements and layout boundaries, not default thresholds.",
        "- Use `decorative_or_case_specific` cases as caution examples when grob-level or polar decoration is hard to generalize.",
        "- Original scripts are provenance. Executable skill recipes must remove hard-coded local paths and expose a clean input schema.",
    ]
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--library", required=True, help="R replica library root")
    parser.add_argument("--archive-root", default=None, help="Skill-relative archive root for copied source scripts")
    parser.add_argument("--out", default=None, help="JSON output path; markdown is written next to it")
    parser.add_argument("--out-md", default=None)
    parser.add_argument("--out-json", default=None)
    args = parser.parse_args()

    library = Path(args.library).expanduser().resolve()
    if not library.exists():
        print(f"Library not found: {library}", file=sys.stderr)
        return 2
    archive_root = Path(args.archive_root) if args.archive_root else None
    scripts = sorted(p for p in library.rglob("*") if p.is_file() and p.suffix.lower() in SCRIPT_EXTS)
    records = []
    for script in scripts:
        case_dir = next((parent for parent in script.parents if parent.parent == library), script.parent)
        records.append(script_metadata(script, case_dir, library, archive_root))

    if args.out:
        out_json = Path(args.out)
        out_md = out_json.with_suffix(".md")
    else:
        if not args.out_json or not args.out_md:
            parser.error("Either --out or both --out-json and --out-md are required")
        out_json = Path(args.out_json)
        out_md = Path(args.out_md)
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps({"library": str(library), "script_count": len(records), "records": records}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    write_markdown(records, out_md, library)
    print(f"Indexed {len(records)} scripts")
    print(f"Wrote {out_json}")
    print(f"Wrote {out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
